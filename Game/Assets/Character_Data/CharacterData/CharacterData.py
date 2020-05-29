from openpyxl import load_workbook
import struct

#Excelからロード
workbook = load_workbook(filename = "C:/Users/sisim/Documents/Hoverboard/Hoverboat/Game/Assets/Character_Data/Character_Data.xlsx",
                         read_only = True)
sheet = workbook["CharaData"]
accel=0.0         #加速度
friction=0.0      #摩擦力
sutearing = 0.0   #ステアリング
charaNoKazu = 0   #キャラの数
rowNo = 5         #列ごとのカウント用の行番号

#キャラの数をカウント
while True:
 if sheet.cell(column = 2, row = rowNo).value == "end":
    #"end"までたどり着くと終わり。
    break
 charaNoKazu += 1
 rowNo += 1

 #キャラの数を抽出(int型)
fp = open("data.ai", "wb")
fp.write(struct.pack("i", charaNoKazu))
rowNo = 5

#Excelからデータを抽出
while True:
 if sheet.cell(column = 2, row = rowNo).value == "end":
     #"end"までたどり着くと終わり。
     break
#加速度を抽出(float型)
 accel = sheet.cell(column = 3, row = rowNo).value
 fp.write(struct.pack("f", accel))

#摩擦度を抽出(float型)
 friction = sheet.cell(column = 4, row = rowNo).value
 fp.write(struct.pack("f", friction))

#ステアリングを抽出(float型)
 sutearing = sheet.cell(column = 5, row = rowNo).value
 fp.write(struct.pack("f", sutearing))

#モデルパスを抽出()
 modelfbx = sheet.cell(column = 6, row = rowNo).value
 fp.write(modelfbx.encode())
 #文字列のバイト数を取得
 strByte = len(modelfbx.encode())
 #残りのバイト数にダミー(0)を入れる。
 dummyData = 0
 fp.write(dummyData.to_bytes(256 - strByte, "little"))
 rowNo +=1

